import pymysql
import pandas as pd
import time
import sys
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import *
from datetime import datetime, timedelta
from holiday_calculator import (
    is_datetime_conversion,
    holiday_is_valid,
    next_day,
    count_holiday,
)
from stock_data import Kiwoom
from openpyxl import load_workbook


today = datetime(2019, 9, 15).strftime("%Y%m%d")
# today = datetime.today().strftime("%Y%m%d")
excel_row = 1701
data_box = []
stock_items = []

db = pymysql.connect(
    host="localhost",
    user="root",
    password="As731585!",
    db="foreign_ins_purchases",
    charset="utf8mb4",
)

cur = db.cursor(pymysql.cursors.DictCursor)
db.commit()
app = QApplication(sys.argv)
kiwoom = Kiwoom()
kiwoom.comm_connect()


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs,
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn"t exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: "/path/to/file.xlsx")
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: "Sheet1")
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # Python 2.x: define [FileNotFoundError] exception if it doesn"t exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    if "index" not in to_excel_kwargs:
        to_excel_kwargs["index"] = False

    try:
        # try to open an existing workbook
        if "header" not in to_excel_kwargs:
            to_excel_kwargs["header"] = True
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
            to_excel_kwargs["header"] = False

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    except FileNotFoundError:
        # file does not exist yet, we will create it
        to_excel_kwargs["header"] = True

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def save_data(data):
    global stock_items
    stock_items = []
    stock_items += data
    return data


def is_db_validate():
    global today
    try:
        sql = f"SELECT * FROM foreign_ins_purchases.`{today}`"
        cur.execute(sql)
        time.sleep(1)
        result_set = cur.fetchall()
        if result_set:
            return save_data(result_set)
    except pymysql.err.ProgrammingError as e:
        print(e)
        today = (is_datetime_conversion(today) - timedelta(days=1)).strftime("%Y%m%d")
        is_db_validate()


def start_price_constrast_fluc_calculator(standard_price, check_prcie):
    result = round(((check_prcie - standard_price) / standard_price) * 100, 2,)
    return result


def time_minus(first, second):
    first = is_datetime_conversion(first)
    second = is_datetime_conversion(second)
    result = first - second
    return result


def is_low_fluctuation_validate(rate_fluctuation):
    if fluctuation >= 3 and rate_3fluctuation < 3 and int(data_date) >= int(next_day):
        return True
    return False


def is_high_fluctuation_validate(rate_fluctuation):
    if fluctuation >= 9 and int(data_date) >= int(next_day):
        return True

    return False


def get_rate_fluctuation(rate_fluctuation):
    if fluctuation >= rate_fluctuation:
        return fluctuation


def get_rate_fluctuation_day(rate_fluctuation):
    if fluctuation >= rate_fluctuation:
        print(f"get_rate_flutuation_day at data_date : {data_date}")
        return data_date


def get_rate_fluctuation_price(rate_fluctuation):
    if fluctuation >= rate_fluctuation:
        return data_high_price


def get_stock_data_day(day):
    count_sub_day = count_holiday(day, 19, False)
    sub_day = is_datetime_conversion(day) + timedelta(days=19)
    while count_sub_day:
        temp_count_sub_day = count_sub_day
        count_sub_day = count_holiday(sub_day, count_sub_day, False)
        sub_day = is_datetime_conversion(sub_day) + timedelta(days=temp_count_sub_day)
    return sub_day


if __name__ == "__main__":
    while int(today) >= 20180323:
        today = (is_datetime_conversion(today) - timedelta(days=1)).strftime("%Y%m%d")
        print(today)
        is_db_validate()

        for stock_item in stock_items:
            (
                name,
                code,
                today_end_price,
                next_day,
                next_day_start_price,
                next_day_end_price,
                next_day_high_fluc,
                rate_3fluctuation,
                rate_9fluctuation,
            ) = (
                stock_item["종목명"],
                stock_item["종목코드"],
                stock_item["당일종가"],
                stock_item["명일"],
                stock_item["명일시가"],
                stock_item["명일종가"],
                stock_item["명일고가등락률"],
                stock_item["명일고가등락률"],
                0,
            )
            three_percent_day = ""
            three_percent_price = ""
            nine_percent_day = ""
            nine_percent_price = ""
            data_assist_day = get_stock_data_day(today).strftime("%Y%m%d")
            print(name, today)

            if "KODEX" not in name and "TIGER 200" not in name and next_day_start_price:
                if isinstance(rate_3fluctuation, str):
                    rate_3fluctuation = 0
                if rate_3fluctuation >= 9:
                    rate_9fluctuation = rate_3fluctuation
                elif next_day_start_price:
                    while rate_3fluctuation < 3 or rate_9fluctuation < 9:
                        if data_assist_day > datetime.today().strftime("%Y%m%d"):
                            data_assist_day = datetime.today().strftime("%Y%m%d")
                            break_sw = True
                        else:
                            break_sw = False
                        kiwoom.on_excel_sw = True
                        kiwoom.set_input_value("종목코드", code)
                        kiwoom.set_input_value("조회일자", data_assist_day)
                        kiwoom.set_input_value("표시구분", 1)
                        time.sleep(3.6)
                        kiwoom.comm_rq_data("opt10086_req", "opt10086", 0, "0101")

                        datas = kiwoom.excel_assist_data
                        reverse_datas = list(reversed(datas))
                        print(reverse_datas)
                        for data in reverse_datas:
                            data_date = data["날짜"]
                            data_high_price = data["고가"]
                            # 등락률
                            fluctuation = start_price_constrast_fluc_calculator(
                                next_day_start_price, data_high_price
                            )

                            # 3% 확인
                            if is_low_fluctuation_validate(3):

                                rate_3fluctuation = fluctuation
                                three_percent_day = data_date
                                three_percent_price = data_high_price

                            # 9% 확인
                            if is_high_fluctuation_validate(9):
                                rate_9fluctuation = fluctuation
                                nine_percent_day = data_date
                                nine_percent_price = data_high_price

                            if rate_3fluctuation >= 3 and rate_9fluctuation >= 9:

                                break  # 3% 등락률 9% 등락률을 다 충족했을 시 break
                        if break_sw:

                            break_sw = False

                            break

                        data_assist_day = get_stock_data_day(data_assist_day).strftime(
                            "%Y%m%d"
                        )
                print("-------------------check---------------")
                print(
                    f"data_date : {data_date} 고가 : {data_high_price} name : {name} rate_3fluctuation :{rate_3fluctuation} fluctuation : {fluctuation} three_percent_day:{three_percent_day} rate_9fluctuation : {rate_9fluctuation} nine_percent_day:{nine_percent_day} nine_percent_price : {nine_percent_price}"
                )
                print("-------------------check---------------")
                data_box = [
                    {
                        "날짜": today,
                        "종목": name,
                        "일일외인순매수": int(stock_item["일일외인순매수"]),
                        "일일기관순매수": int(stock_item["일일기관순매수"]),
                        "일일외인기관순매수": int(stock_item["일일외인순매수"])
                        + int(stock_item["일일기관순매수"]),
                        "당일시가": stock_item["당일시가"],
                        "당일전일대비등락률": stock_item["당일전일대비등락률"],
                        "명일전일대비등락률": stock_item["명일전일대비등락률"],
                        "명일고가등락률": next_day_high_fluc,
                        "명일": next_day,
                        "명일시가": next_day_start_price,
                        "종가대비명일시가등락률": start_price_constrast_fluc_calculator(
                            today_end_price, next_day_start_price
                        ),
                        "up3%날짜": three_percent_day if three_percent_day else next_day,
                        "up3%걸린시간": time_minus(three_percent_day, next_day)
                        if three_percent_day
                        else 0,
                        "up3%perscent": rate_3fluctuation
                        if rate_3fluctuation >= 3
                        else "",
                        "up3%price": str(three_percent_price)
                        if three_percent_price
                        else str(next_day_end_price)
                        if rate_3fluctuation >= 3
                        else "",
                        "up9%날짜": nine_percent_day
                        if nine_percent_day
                        else next_day
                        if rate_3fluctuation >= 9
                        else "",
                        "up9%걸린시간": time_minus(nine_percent_day, next_day)
                        if nine_percent_day
                        else 0
                        if rate_3fluctuation >= 9
                        else "",
                        "up9%percent": rate_9fluctuation
                        if rate_9fluctuation
                        else rate_3fluctuation
                        if rate_3fluctuation >= 9
                        else "",
                        "up9%price": nine_percent_price
                        if nine_percent_price
                        else next_day_end_price
                        if rate_3fluctuation >= 9
                        else "",
                    }
                ]

                df = pd.DataFrame(
                    data_box,
                    columns=[
                        "날짜",
                        "종목",
                        "일일외인순매수",
                        "일일기관순매수",
                        "일일외인기관순매수",
                        "당일시가",
                        "당일전일대비등락률",
                        "명일전일대비등락률",
                        "명일고가등락률",
                        "명일",
                        "명일시가",
                        "종가대비명일시가등락률",
                        "up3%날짜",
                        "up3%걸린시간",
                        "up3%perscent",
                        "up3%price",
                        "up9%날짜",
                        "up9%걸린시간",
                        "up9%percent",
                        "up9%price",
                    ],
                )

                append_df_to_excel(
                    "./기록용/real_v3.xlsx",
                    df,
                    "Sheet1",
                    startrow=excel_row,
                    truncate_sheet=False,
                    header=None,
                )
                print(excel_row)
                excel_row += 1

